import os
from datetime import datetime, timezone
from functools import wraps
from io import BytesIO

from dotenv import load_dotenv
from flask import (
    Flask,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from flask_sqlalchemy import SQLAlchemy
from flask_wtf import CSRFProtect
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from werkzeug.security import check_password_hash

load_dotenv()

app = Flask(__name__)

app.config["SECRET_KEY"] = os.getenv("SECRET_KEY", "change-this-before-production")
app.config["SQLALCHEMY_DATABASE_URI"] = os.getenv(
    "DATABASE_URL", "sqlite:///messages.db"
)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["MAX_CONTENT_LENGTH"] = 32 * 1024  # 32 KB request limit
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = (
    os.getenv("SESSION_COOKIE_SECURE", "false").lower() == "true"
)

db = SQLAlchemy(app)
csrf = CSRFProtect(app)


class AnonymousMessage(db.Model):
    __tablename__ = "anonymous_messages"

    id = db.Column(db.Integer, primary_key=True)
    body = db.Column(db.Text, nullable=False)
    created_at = db.Column(
        db.DateTime(timezone=True),
        nullable=False,
        default=lambda: datetime.now(timezone.utc),
    )

    def __repr__(self):
        return f"<AnonymousMessage {self.id}>"


def excel_safe_text(value):
    """Prevent untrusted message text from becoming a spreadsheet formula."""
    if value and value[0] in ("=", "+", "-", "@"):
        return "'" + value
    return value


def admin_required(view_function):
    @wraps(view_function)
    def wrapped(*args, **kwargs):
        if not session.get("is_admin"):
            flash("Please sign in as admin.", "warning")
            return redirect(url_for("admin_login", next=request.path))
        return view_function(*args, **kwargs)

    return wrapped


@app.context_processor
def inject_globals():
    return {
        "site_name": os.getenv("SITE_NAME", "WhisperBox"),
        "current_year": datetime.now().year,
    }


@app.route("/", methods=["GET", "POST"])
def home():
    if request.method == "POST":
        # Honeypot field: normal visitors never fill this.
        # It helps block simple bots without collecting IP addresses.
        if request.form.get("website", "").strip():
            return redirect(url_for("message_sent"))

        message = request.form.get("message", "").strip()

        if not message:
            flash("Please write a message before sending.", "danger")
            return render_template("index.html", message=message)

        if len(message) < 2:
            flash("The message is too short.", "danger")
            return render_template("index.html", message=message)

        if len(message) > 5000:
            flash("Please keep the message under 5,000 characters.", "danger")
            return render_template("index.html", message=message)

        db.session.add(AnonymousMessage(body=message))
        db.session.commit()

        return redirect(url_for("message_sent"))

    return render_template("index.html")


@app.route("/sent")
def message_sent():
    return render_template("sent.html")


@app.route("/admin/login", methods=["GET", "POST"])
def admin_login():
    if session.get("is_admin"):
        return redirect(url_for("admin_dashboard"))

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")

        expected_username = os.getenv("ADMIN_USERNAME", "admin")
        password_hash = os.getenv("ADMIN_PASSWORD_HASH", "")

        if not password_hash:
            flash(
                "Admin password is not configured. Set ADMIN_PASSWORD_HASH in .env.",
                "danger",
            )
            return render_template("admin_login.html")

        if username == expected_username and check_password_hash(
            password_hash, password
        ):
            session.clear()
            session["is_admin"] = True
            session["admin_username"] = username
            flash("Signed in successfully.", "success")

            next_url = request.args.get("next")
            if next_url and next_url.startswith("/"):
                return redirect(next_url)
            return redirect(url_for("admin_dashboard"))

        flash("Invalid username or password.", "danger")

    return render_template("admin_login.html")


@app.route("/admin/logout", methods=["POST"])
@admin_required
def admin_logout():
    session.clear()
    flash("You have been signed out.", "success")
    return redirect(url_for("admin_login"))


@app.route("/admin")
@admin_required
def admin_dashboard():
    page = request.args.get("page", 1, type=int)
    per_page = 20

    pagination = (
        AnonymousMessage.query
        .order_by(AnonymousMessage.created_at.desc(), AnonymousMessage.id.desc())
        .paginate(page=page, per_page=per_page, error_out=False)
    )

    total_messages = AnonymousMessage.query.count()

    return render_template(
        "admin_dashboard.html",
        messages=pagination.items,
        pagination=pagination,
        total_messages=total_messages,
    )


@app.route("/admin/message/<int:message_id>/delete", methods=["POST"])
@admin_required
def delete_message(message_id):
    message = db.get_or_404(AnonymousMessage, message_id)
    db.session.delete(message)
    db.session.commit()
    flash("Message deleted.", "success")
    return redirect(request.referrer or url_for("admin_dashboard"))


@app.route("/admin/export")
@admin_required
def export_messages():
    messages = (
        AnonymousMessage.query
        .order_by(AnonymousMessage.created_at.asc(), AnonymousMessage.id.asc())
        .all()
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Anonymous Messages"

    headers = ["ID", "Message", "Received At (UTC)"]
    sheet.append(headers)

    header_fill = PatternFill("solid", fgColor="172033")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    for message in messages:
        received = message.created_at
        if received and received.tzinfo is not None:
            received = received.astimezone(timezone.utc).replace(tzinfo=None)

        sheet.append([
            message.id,
            excel_safe_text(message.body),
            received.strftime("%Y-%m-%d %H:%M:%S") if received else "",
        ])

    sheet.column_dimensions["A"].width = 10
    sheet.column_dimensions["B"].width = 90
    sheet.column_dimensions["C"].width = 24

    for row in sheet.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        row[2].alignment = Alignment(vertical="top")

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    filename = f"anonymous_messages_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )


@app.errorhandler(413)
def request_too_large(_error):
    flash("The submitted request was too large.", "danger")
    return redirect(url_for("home"))


with app.app_context():
    db.create_all()


if __name__ == "__main__":
    app.run(
        host=os.getenv("FLASK_HOST", "0.0.0.0"),
        port=int(os.getenv("PORT", os.getenv("FLASK_PORT", "5000"))),
        debug=os.getenv("FLASK_DEBUG", "false").lower() == "true",
    )
